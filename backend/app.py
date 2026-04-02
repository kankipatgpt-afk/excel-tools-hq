from dotenv import load_dotenv
from supabase import create_client
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from sqlalchemy import create_engine, text
import pandas as pd
import os
import uuid
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

load_dotenv()

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 10 * 1024 * 1024  # 10 MB limit
CORS(app, resources={r"/*": {"origins": "*"}})

DATABASE_URL = os.getenv("DATABASE_URL")
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY")
SUPABASE_BUCKET = os.getenv("SUPABASE_BUCKET", "tool-files")

if not DATABASE_URL:
    raise ValueError("DATABASE_URL is missing")

engine = create_engine(DATABASE_URL, pool_pre_ping=True)

supabase = None

def get_supabase():
    global supabase
    if supabase is None:
        supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)
    return supabase


def init_db():
    try:
        with engine.connect() as conn:
            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS leads (
                    id SERIAL PRIMARY KEY,
                    name TEXT,
                    email TEXT,
                    message TEXT
                )
            """))

            conn.execute(text("""
                CREATE TABLE IF NOT EXISTS tool_history (
                    id SERIAL PRIMARY KEY,
                    tool_name TEXT,
                    original_file TEXT,
                    output_file TEXT,
                    user_email TEXT,
                    user_name TEXT,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """))

            conn.commit()
    except Exception as e:
        print("DB init skipped:", e)

init_db()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
OUTPUT_FOLDER = os.path.join(BASE_DIR, "outputs")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def upload_file_to_supabase(local_path, storage_path, content_type="application/octet-stream"):
    with open(local_path, "rb") as f:
        file_bytes = f.read()

    client = get_supabase()

    client.storage.from_(SUPABASE_BUCKET).upload(
        path=storage_path,
        file=file_bytes,
        file_options={"content-type": content_type, "upsert": "true"},
    )

    signed = client.storage.from_(SUPABASE_BUCKET).create_signed_url(storage_path, 3600)
    return signed.get("signedURL") or signed.get("signed_url")


@app.route('/')
def home():
    return "Flask backend running"

@app.route('/submit', methods=['POST'])
def submit():
    data = request.json

    name = data.get('name')
    email = data.get('email')
    message = data.get('message')

    with engine.connect() as conn:
        conn.execute(
            text("INSERT INTO leads (name, email, message) VALUES (:name, :email, :message)"),
            {"name": name, "email": email, "message": message}
        )
        conn.commit()

    return jsonify({"message": "Saved to PostgreSQL!"})


def clean_extra_spaces(value):
    if isinstance(value, str):
        # remove leading/trailing spaces and reduce multiple spaces to single
        return " ".join(value.split())
    return value


@app.route('/trim-spaces', methods=['POST'])
def trim_spaces():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files['file']

    user_email = request.form.get("user_email", "")
    user_name = request.form.get("user_name", "")

    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in ['.xlsx', '.xls']:
        return jsonify({"error": "Only Excel files are allowed"}), 400

    unique_name = f"{uuid.uuid4()}_{file.filename}"
    input_path = os.path.join(UPLOAD_FOLDER, unique_name)

    output_name = f"trimmed_{file.filename}"
    output_path = os.path.join(OUTPUT_FOLDER, output_name)

    file.save(input_path)

    try:
        # Read all sheets
        excel_data = pd.read_excel(input_path, sheet_name=None)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:      
            for sheet_name, df in excel_data.items():
                cleaned_df = df.apply(lambda col: col.map(clean_extra_spaces))
                cleaned_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Save tool history
        print("Step 1: Excel cleaned successfully")

        print("Step 2: Saving tool history to database")
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            for sheet_name, df in excel_data.items():
                cleaned_df = df.apply(lambda col: col.map(clean_extra_spaces))
                cleaned_df.to_excel(writer, sheet_name=sheet_name, index=False)

        print("Step 1: Excel cleaned successfully")

        print("Step 2: Saving tool history to database")
        with engine.connect() as conn:
            conn.execute(
                text("""
                    INSERT INTO tool_history (tool_name, original_file, output_file, user_email, user_name)
                    VALUES (:tool_name, :original_file, :output_file, :user_email, :user_name)
                """),
                {
                    "tool_name": "Trim Spaces",
                    "original_file": file.filename,
                    "output_file": output_name,
                    "user_email": user_email,
                    "user_name": user_name
                }
            )
            conn.commit()
        print("Step 3: Database save successful")

        print("Step 4: Uploading output file to Supabase")
        storage_path = f"outputs/{output_name}"
        download_url = upload_file_to_supabase(
            output_path,
            storage_path,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        print("Step 5: Supabase upload successful")

        return jsonify({
            "message": "File processed successfully",
            "download_url": download_url
        })

    except Exception as e:
        print("trim_spaces error:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/excel-metadata', methods=['POST'])
def excel_metadata():
    if 'file' not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    file = request.files['file']

    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400

    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in ['.xlsx', '.xls']:
        return jsonify({"error": "Only Excel files are allowed"}), 400

    unique_name = f"{uuid.uuid4()}_{file.filename}"
    input_path = os.path.join(UPLOAD_FOLDER, unique_name)
    file.save(input_path)

    try:
        excel_data = pd.read_excel(input_path, sheet_name=None)

        metadata = {
            "temp_file": unique_name,
            "sheets": []
        }

        for sheet_name, df in excel_data.items():
            metadata["sheets"].append({
                "sheet_name": sheet_name,
                "columns": list(df.columns.astype(str))
            })

        return jsonify(metadata)

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@app.route('/remove-duplicates', methods=['POST'])
def remove_duplicates():
    temp_file = request.form.get("temp_file")
    sheet_name = request.form.get("sheet_name")
    action_type = request.form.get("action_type")
    mode = request.form.get("mode")
    selected_column = request.form.get("selected_column")
    user_email = request.form.get("user_email", "")
    user_name = request.form.get("user_name", "")

    if not temp_file:
        return jsonify({"error": "Missing temp file"}), 400

    if not sheet_name:
        return jsonify({"error": "Sheet name is required"}), 400

    if action_type not in ["highlight", "delete"]:
        return jsonify({"error": "Invalid action type"}), 400

    if mode not in ["row", "column"]:
        return jsonify({"error": "Invalid mode"}), 400

    input_path = os.path.join(UPLOAD_FOLDER, temp_file)

    if not os.path.exists(input_path):
        return jsonify({"error": "Uploaded temp file not found"}), 400

    original_name = temp_file.split("_", 1)[1] if "_" in temp_file else temp_file
    output_name = f"dedup_{original_name}"
    output_path = os.path.join(OUTPUT_FOLDER, output_name)

    try:
        excel_data = pd.read_excel(input_path, sheet_name=None)

        if sheet_name not in excel_data:
            return jsonify({"error": "Selected sheet not found"}), 400

        target_df = excel_data[sheet_name].copy()

        if mode == "row":
            duplicate_mask = target_df.duplicated(keep=False)

            if action_type == "delete":
                target_df = target_df.drop_duplicates()
                excel_data[sheet_name] = target_df

                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    for s_name, df in excel_data.items():
                        df.to_excel(writer, sheet_name=s_name, index=False)

            elif action_type == "highlight":
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    for s_name, df in excel_data.items():
                        df.to_excel(writer, sheet_name=s_name, index=False)

                wb = load_workbook(output_path)
                ws = wb[sheet_name]

                yellow_fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")

                for idx, is_dup in enumerate(duplicate_mask.tolist(), start=2):
                    if is_dup:
                        for col_idx in range(1, ws.max_column + 1):
                            ws.cell(row=idx, column=col_idx).fill = yellow_fill

                wb.save(output_path)

        elif mode == "column":
            if not selected_column:
                return jsonify({"error": "Please select a column"}), 400

            if selected_column not in target_df.columns.astype(str).tolist():
                return jsonify({"error": "Selected column not found"}), 400

            actual_col = None
            for col in target_df.columns:
                if str(col) == selected_column:
                    actual_col = col
                    break

            duplicate_mask = target_df[actual_col].duplicated(keep=False)

            if action_type == "delete":
                target_df = target_df.drop_duplicates(subset=[actual_col])
                excel_data[sheet_name] = target_df

                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    for s_name, df in excel_data.items():
                        df.to_excel(writer, sheet_name=s_name, index=False)

            elif action_type == "highlight":
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    for s_name, df in excel_data.items():
                        df.to_excel(writer, sheet_name=s_name, index=False)

                wb = load_workbook(output_path)
                ws = wb[sheet_name]

                yellow_fill = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")

                selected_excel_col = None
                for col_idx in range(1, ws.max_column + 1):
                    if str(ws.cell(row=1, column=col_idx).value) == selected_column:
                        selected_excel_col = col_idx
                        break

                for idx, is_dup in enumerate(duplicate_mask.tolist(), start=2):
                    if is_dup and selected_excel_col:
                        ws.cell(row=idx, column=selected_excel_col).fill = yellow_fill

                wb.save(output_path)

        with engine.connect() as conn:
            conn.execute(
                text("""
                    INSERT INTO tool_history (tool_name, original_file, output_file, user_email, user_name)
                    VALUES (:tool_name, :original_file, :output_file, :user_email, :user_name)
                """),
                {
                    "tool_name": "Remove Duplicates",
                    "original_file": temp_file,
                    "output_file": output_name,
                    "user_email": user_email,
                    "user_name": user_name
                }
            )
            conn.commit()

        storage_path = f"outputs/{output_name}"
        download_url = upload_file_to_supabase(
            output_path,
            storage_path,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        return jsonify({
            "message": "File processed successfully",
            "download_url": download_url
        })

    except Exception as e:
        print("remove_duplicates error:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/tool-history', methods=['GET'])
def tool_history():
    try:
        with engine.connect() as conn:
            result = conn.execute(text("""
                SELECT id, tool_name, original_file, output_file, created_at, user_email, user_name
                FROM tool_history
                ORDER BY id DESC
            """))

            rows = result.fetchall()

            history = []
            for row in rows:
                history.append({
                    "id": row[0],
                    "tool_name": row[1],
                    "original_file": row[2],
                    "output_file": row[3],
                    "created_at": str(row[4]) if row[4] else "",
                    "user_email": row[5] if row[5] else "",
                    "user_name": row[6] if row[6] else ""
                })

        return jsonify(history)

    except Exception as e:
        print("Tool history error:", str(e))
        return jsonify({"error": str(e)}), 500



@app.route('/download-output/<filename>', methods=['GET'])
def download_output(filename):
    try:
        return send_from_directory(OUTPUT_FOLDER, filename, as_attachment=True)
    except Exception as e:
        return jsonify({"error": str(e)}), 404

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)